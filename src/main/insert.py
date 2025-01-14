from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
import json
from collections import Counter
import platform

# Configuração de print
def lista_para_string(lista):
    """Converte uma lista em uma string separada por vírgulas."""
    return ''.join(map(str, lista)) if lista else ""


def to_int(value):
    """Converte um valor para inteiro. Retorna 0 se a conversão falhar."""
    try:
        return float(float(value))  # Para lidar com valores como "12.0"
    except (ValueError, TypeError):
        return 0
        

def converter_data(data_str):
    """Converte uma string para um objeto datetime, aceitando múltiplos formatos."""
    formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]
    for fmt in formatos:
        try:
            return datetime.strptime(data_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Formato de data inválido: {data_str}")

def validar_dados(formulario):
    """Valida os dados do formulário antes de processar."""
    campos_obrigatorios = ["missao", "inicio_miss", "term_miss"]
    for campo in campos_obrigatorios:
        if not formulario.get(campo):
            raise ValueError(f"O campo obrigatório '{campo}' está ausente ou vazio.")

def criar_estilo_data(workbook):
    """Cria um estilo para formatação de datas no formato brasileiro."""
    if "br_date" not in workbook.named_styles:
        estilo_data = NamedStyle(name="br_date")
        estilo_data.number_format = "DD/MM/YYYY"
        workbook.add_named_style(estilo_data)

def aplicar_formato_data(celula):
    """Aplica o formato de data brasileira em uma célula."""
    celula.number_format = "DD/MM/YYYY"

def verificar_arquivo(file_path):
    """Verifica se o arquivo Excel existe no diretório especificado."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

def remover_protecao_exibicao(file_path):
    """Remove a proteção de exibição protegida do Excel marcada pelo Windows."""
    if platform.system() == "Windows":
        try:
            result = os.system(f'powershell -Command "Unblock-File -Path \'{file_path}\'"')
            if result != 0:
                print.warning("Falha ao remover proteção de exibição protegida. Verifique permissões.")
        except Exception as e:
            print.error(f"Erro ao tentar remover proteção: {e}")
    else:
        print.info("Remoção de proteção de exibição protegida não é necessária neste SO.")

def find_next_empty_row(sheet, start_row, column):
    """Encontra a próxima linha vazia em uma coluna específica."""
    for row in range(start_row, sheet.max_row + 2):
        if sheet.cell(row=row, column=column).value is None:
            return row
    return sheet.max_row + 1

def inserir_dados(file_path, formulario):
    """Insere os dados do formulário na planilha do Excel."""
    try:
        # Validar dados
        validar_dados(formulario)

        # Verificar se o arquivo existe
        verificar_arquivo(file_path)

        # Carregar a planilha
        wb = load_workbook(file_path)
        ws = wb.active

        # Criar estilo de data
        criar_estilo_data(wb)

        # Encontra a próxima linha vazia
        next_empty_row = find_next_empty_row(ws, 4, 4)

        # Validar e converter datas para objetos datetime
        campos_data = ["inicio_miss", "term_miss", "inicio_desl", "term_desl"]
        datas_convertidas = {campo: converter_data(formulario[campo]) if formulario.get(campo) else None for campo in campos_data}

        # Inserir dados do formulário
        campos_texto = [
            ("missao", 4),
            ("evento", 5),
            ("objetivo_missao", 6),
            ("justificativa", 7),
            ("eventual_prej", 9),
            ("ambito", 10),
        ]

        for campo, coluna in campos_texto:
            valor = lista_para_string(formulario.get(campo, '')).strip() 
            ws.cell(row=next_empty_row, column=coluna, value=valor)
        
        # Inserir os cargos diretamente
        cargos = formulario.get("cargoParticipantes", [])
        cargos_count = Counter(cargos)
        cargos_str = ', '.join([f"{count} {cargo}" for cargo, count in cargos_count.items()]) 
        ws.cell(row=next_empty_row, column=19, value=cargos_str if cargos else "Nenhum cargo selecionado")
    
        #Obter os dados selecionadas através do checkbox

        cidadesSelecionadas = formulario.get("cidadesSelecionadas", [])

        # Inserir cidades selecionadas
        cidades_str = ', '.join(cidadesSelecionadas)
        # Validar e converter cidades/estados para strings
        
        # Inserir datas convertidas

        # Inicializar listas para armazenar cidades e estados separados
        lista_cidades = []
        lista_estados = []

        # Para cada checkbox (cidade/estado), separar cidade e estado
        for cidade_estado in cidadesSelecionadas:
            try:
                if " - " in cidade_estado:  # Verifica se o separador está presente
                    cidade, estado = cidade_estado.split(" - ", 1)  # Divide apenas uma vez
                    lista_cidades.append(cidade)  # Adiciona a cidade na lista
                    lista_estados.append(estado)  # Adiciona o estado na lista
                else:
                    raise ValueError("Formato inválido")  # Força um erro se o formato for inválido
            except ValueError:
                lista_cidades.append("Formato de cidade inválido")
                lista_estados.append("Formato de estado inválido")

        # Unir todas as cidades e estados em strings separadas, separados por vírgula
        cidades_str = ', '.join(lista_cidades) if lista_cidades else "Nenhuma cidade selecionada"
        estados_str = ', '.join(lista_estados) if lista_estados else "Nenhum estado selecionado"

        # Inserção no Excel
        ws.cell(row=next_empty_row, column=11, value=estados_str)  # Coluna 13: Todos os estados
        ws.cell(row=next_empty_row, column=12, value=cidades_str)  # Coluna 12: Todas as cidades

        # Exibir os valores separados (para depuração)

        # Inserir dados numéricos
        campos_numericos = [
            ("prioridade", 8),
            ("num_diarias",17),
            ("numeroParticipantes", 18),
            ("preco_diarias", 20),
            ("preco_pass", 21),
        ]

        for campo, coluna in campos_numericos:
            ws.cell(row=next_empty_row, column=coluna, value=to_int(formulario.get(campo, 0)))

        # Inserir datas formatadas
        colunas_datas = [13, 14, 15, 16]
        for campo, coluna in zip(campos_data, colunas_datas):
            if datas_convertidas[campo]:
                cell = ws.cell(row=next_empty_row, column=coluna, value=datas_convertidas[campo])
                aplicar_formato_data(cell)

        # Salvar alterações
        wb.save(file_path)
        
        print(f"true")


    except Exception as e:
        ...
if __name__ == "__main__":
    try:
        sys.stdin.reconfigure(encoding='utf-8')
        input_data = sys.stdin.read()
        formulario = json.loads(input_data)

        file_path = os.getenv("PLANILHA_PATH", r"C:\Users\anderson.teles\Desktop\projetos-amazonia-azul\feat_sistema_diarias\src\main\PROGRAMAÇÃO NACIONAL DE DIÁRIAS E PASSAGENS - 2025 - CGMAB.xlsx")

        remover_protecao_exibicao(file_path)
        inserir_dados(file_path, formulario)
    except:
        ...